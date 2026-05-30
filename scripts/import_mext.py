#!/usr/bin/env python3
from __future__ import annotations
"""
import_mext.py — 日本食品標準成分表（文部科学省）→ Atom JSON 変換スクリプト

【準備】
1. 文部科学省の公式 Excel を取得
   python3 scripts/import_mext.py --download-latest

2. すでに Excel を持っている場合
   python3 scripts/import_mext.py <ダウンロードしたExcelファイルのパス>

   例:
   python3 scripts/import_mext.py ~/Downloads/mext_01110.xlsx

【出力】
   data/seed-atoms/mext-atoms.json
"""

import json
import argparse
import re
import sys
import tempfile
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).parent.parent
OUT_FILE = ROOT / "data" / "seed-atoms" / "mext-atoms.json"
MEXT_PAGE_URL = "https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html"
MEXT_SOURCE_NAME = "日本食品標準成分表（八訂）増補2023年"
MEXT_LICENSE_NOTE = "出典表記: 日本食品標準成分表（八訂）増補2023年から引用（又は出典）"

# ── 食品群コード → category マッピング ───────────────────────────────────────
FOOD_GROUP_MAP = {
    "01": "grain",
    "02": "starch_root",
    "03": "sweetener",
    "04": "legume",
    "05": "nut_seed",
    "06": "vegetable",
    "07": "fruit",
    "08": "mushroom",
    "09": "seaweed",
    "10": "seafood",
    "11": "meat",
    "12": "egg",
    "13": "dairy",
    "14": "fat_oil",
    "15": "confectionery",
    "16": "beverage",
    "17": "seasoning",
    "18": "processed_food",
}

# ── アレルゲン・リスクタグ自動付与ルール ──────────────────────────────────────
ALLERGEN_RULES = [
    (r"小麦|グルテン|うどん|パン|ラーメン|マカロニ|スパゲ",    ["wheat_allergen", "gluten_related"]),
    (r"牛乳|乳|チーズ|バター|ヨーグルト|クリーム|ホエイ",      ["milk_allergen"]),
    (r"大豆|豆腐|味噌|醤油|納豆|豆乳|枝豆",                    ["soy_allergen"]),
    (r"卵|たまご|マヨネーズ",                                   ["egg_allergen"]),
    (r"えび|海老|蝦",                                           ["shellfish_allergen"]),
    (r"かに|蟹",                                                ["shellfish_allergen"]),
    (r"ピーナッツ|落花生",                                      ["peanut_allergen"]),
    (r"くるみ|アーモンド|カシューナッツ|マカダミア|ヘーゼル",   ["nut_allergen"]),
    (r"ごま|胡麻",                                              ["sesame_allergen"]),
    (r"さば|鯖",                                                ["fish_allergen"]),
    (r"いくら|さけ|鮭|たら|鱈|まぐろ|鮪|あじ|鯵",              ["fish_allergen"]),
    (r"カフェイン|コーヒー|紅茶|ウーロン",                      ["high_caffeine"]),
]

# ── 栄養値 → known_actions 自動付与ルール ────────────────────────────────────
# (栄養素キー, 閾値per100g, action)
NUTRITION_ACTIONS = [
    ("protein_g",     15.0,  "protein_source"),
    ("fat_g",         20.0,  "fat_source"),
    ("fiber_g",        4.0,  "fiber_source"),
    ("calcium_mg",   200.0,  "calcium_source"),
    ("iron_mg",        2.0,  "iron_source"),
    ("zinc_mg",        2.0,  "zinc_source"),
    ("magnesium_mg",  50.0,  "magnesium_source"),
    ("vitamin_c_mg",  30.0,  "vitamin_c_source"),
    ("vitamin_d_ug",   2.0,  "vitamin_d_source"),
    ("vitamin_b12_ug", 1.0,  "vitamin_b12_source"),
]

# 発酵食品キーワード → 追加アクション
FERMENTED_KEYWORDS = r"みそ|味噌|醤油|しょうゆ|納豆|漬|ヨーグルト|チーズ|酢|ぬか|キムチ|テンペ|塩こうじ|酒粕"


def _safe_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    if s in ("", "-", "0)", "Tr", "tr", "(0)", "(Tr)"):
        return 0.0
    try:
        return float(re.sub(r"[()]", "", s))
    except ValueError:
        return None


def _detect_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    データ開始行と列インデックスを検出する。
    食品番号が '01001' のような 5 桁数字のセルを探す。
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(v).strip() if v is not None else "" for v in row[:4]]
        if any(re.match(r"^\d{5}$", cell) for cell in cells):
            # この行がデータの最初の行
            # ヘッダーはその前の行を探す
            header_row_idx = row_idx - 1
            break
    else:
        raise ValueError(
            "食品番号（例：01001）が見つかりませんでした。\n"
            "ファイルが正しい MEXT 成分表 Excel か確認してください。"
        )

    # ヘッダー行の列名を取得
    col_map: dict[str, int] = {}
    for hr in range(max(1, header_row_idx - 2), header_row_idx + 1):
        for col_idx, cell in enumerate(ws[hr]):
            val = str(cell.value).strip() if cell.value is not None else ""
            if val:
                col_map[val] = col_idx

    return row_idx - 1, col_map  # (data_start_row_1indexed, col_map)


def _find_col(col_map: dict[str, int], *candidates) -> int | None:
    for c in candidates:
        if c in col_map:
            return col_map[c]
    return None


def _build_atom(row: tuple, col: dict[str, int | None]) -> dict | None:
    def g(key):
        idx = col.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    food_no  = str(g("食品番号") or "").strip()
    name_ja  = str(g("食品名") or "").strip()

    if not food_no or not re.match(r"^\d{5}$", food_no):
        return None
    if not name_ja or name_ja in ("-", "－"):
        return None

    group_code = food_no[:2]
    category = FOOD_GROUP_MAP.get(group_code, "other")

    # 栄養素
    nutrition = {
        "energy_kcal":   _safe_float(g("エネルギー_kcal")),
        "water_g":       _safe_float(g("水分")),
        "protein_g":     _safe_float(g("たんぱく質")),
        "fat_g":         _safe_float(g("脂質")),
        "carb_g":        _safe_float(g("炭水化物")),
        "fiber_g":       _safe_float(g("食物繊維総量")),
        "ash_g":         _safe_float(g("灰分")),
        "sodium_mg":     _safe_float(g("ナトリウム")),
        "potassium_mg":  _safe_float(g("カリウム")),
        "calcium_mg":    _safe_float(g("カルシウム")),
        "magnesium_mg":  _safe_float(g("マグネシウム")),
        "phosphorus_mg": _safe_float(g("リン")),
        "iron_mg":       _safe_float(g("鉄")),
        "zinc_mg":       _safe_float(g("亜鉛")),
        "vitamin_c_mg":  _safe_float(g("ビタミンC")),
        "vitamin_d_ug":  _safe_float(g("ビタミンD")),
        "vitamin_b12_ug":_safe_float(g("ビタミンB12")),
    }
    nutrition = {k: v for k, v in nutrition.items() if v is not None}

    # known_actions を栄養値から自動生成
    known_actions = []
    for nutr_key, threshold, action in NUTRITION_ACTIONS:
        val = nutrition.get(nutr_key)
        if val is not None and val >= threshold:
            known_actions.append(action)

    # 発酵食品
    if re.search(FERMENTED_KEYWORDS, name_ja):
        known_actions.append("fermented_food")
        known_actions.append("probiotic_potential")

    # risk_tags をアレルゲンルールから自動生成
    risk_tags = []
    for pattern, tags in ALLERGEN_RULES:
        if re.search(pattern, name_ja):
            for t in tags:
                if t not in risk_tags:
                    risk_tags.append(t)

    # 高エネルギー脂質
    if nutrition.get("fat_g", 0) >= 50:
        if "dose_review_required" not in risk_tags:
            risk_tags.append("dose_review_required")

    source_url = MEXT_PAGE_URL
    retrieved_at = datetime.now(timezone.utc).isoformat()

    atom = {
        "atom_id":   f"mext_{food_no}",
        "name_ja":   name_ja,
        "name_en":   name_ja,   # 英語名は MEXT データに含まれないため暫定
        "atom_type": "ingredient",
        "category":  category,
        "domain":    "food_bio",
        "source_databases": ["MEXT"],
        "canonical_ids": {
            "mext_food_id": food_no,
        },
        "source": {
            "name":       MEXT_SOURCE_NAME,
            "food_no":    food_no,
            "url":        source_url,
        },
        "mext": {
            "canonical_id":     f"mext:{food_no}",
            "source_id":        food_no,
            "source_name":      MEXT_SOURCE_NAME,
            "source_url":       source_url,
            "source_priority":  1,
            "license_note":     MEXT_LICENSE_NOTE,
            "retrieved_at":     retrieved_at,
            "nutrients":        nutrition,
        },
        "nutrition_per_100g": nutrition,
        "known_actions":  list(dict.fromkeys(known_actions)),
        "possible_bonds": _derive_bonds(known_actions, category),
        "risk_tags":      risk_tags,
        "evidence_keywords": [name_ja],
        "regulatory_notes":  [],
        "process_keywords":  [],
        "supplier_keywords": [],
    }
    return atom


def _find_latest_excel_url() -> str:
    req = urllib.request.Request(MEXT_PAGE_URL, headers={"User-Agent": "FormulaFuseStudio/1.0"})
    with urllib.request.urlopen(req, timeout=20) as resp:
        html = resp.read().decode("utf-8", errors="replace")

    candidates = re.findall(r'href="([^"]+\.xls[xm]?)"', html, flags=re.IGNORECASE)
    if not candidates:
        raise RuntimeError("MEXT ページから Excel リンクを検出できませんでした。")

    # 第2章（データ）の本表 Excel を最優先。ページ上では最初の Excel が本表。
    url = candidates[0]
    return urllib.parse.urljoin(MEXT_PAGE_URL, url)


def download_latest_excel(download_dir: Path | None = None) -> Path:
    url = _find_latest_excel_url()
    out_dir = download_dir or Path(tempfile.gettempdir())
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = Path(urllib.parse.urlparse(url).path).name or "mext-latest.xlsx"
    out_path = out_dir / filename

    print(f"⬇️  MEXT Excel を取得: {url}")
    req = urllib.request.Request(url, headers={"User-Agent": "FormulaFuseStudio/1.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        out_path.write_bytes(resp.read())
    print(f"💾 ダウンロード: {out_path}")
    return out_path


def _derive_bonds(actions: list[str], category: str) -> list[str]:
    bonds = []
    if "protein_source" in actions:
        bonds.extend(["protein_degradation", "hydrolysis"])
    if "fiber_source" in actions:
        bonds.append("prebiotic_effect")
    if "fermented_food" in actions:
        bonds.append("fermentation")
    if category in ("grain", "starch_root"):
        bonds.append("starch_degradation")
    if category in ("dairy",):
        bonds.extend(["fermentation", "acidification"])
    if category in ("legume",):
        bonds.extend(["fermentation", "hydrolysis"])
    return list(dict.fromkeys(bonds))


def _build_col_map_from_headers(ws, data_start_row: int) -> dict[str, int | None]:
    """
    ヘッダー行群を結合して列マップを作る。
    MEXT Excel は複数行ヘッダーなので、セルが空の場合は上の行から引き継ぐ。
    """
    def norm(value: str) -> str:
        return re.sub(r"\s+", "", value.replace("\u3000", ""))

    # MEXT Excel は 10 行前後の複数行ヘッダーを持つため、データ開始直前までを結合する
    header_rows = []
    for r in range(1, data_start_row):
        header_rows.append([str(c.value or "").strip() for c in ws[r]])

    # 列ごとに "有効な名前" を結合
    max_cols = max(len(r) for r in header_rows)
    col_names: list[str] = []
    for ci in range(max_cols):
        parts = []
        for hr in header_rows:
            v = hr[ci] if ci < len(hr) else ""
            if v and v not in parts:
                parts.append(v)
        col_names.append("_".join(parts) if parts else "")

    # 既知の列名を正規化してマップを作る
    col_map: dict[str, int | None] = {
        "食品番号":     None,
        "食品名":       None,
        "エネルギー_kcal": None,
        "水分":         None,
        "たんぱく質":   None,
        "脂質":         None,
        "炭水化物":     None,
        "食物繊維総量": None,
        "灰分":         None,
        "ナトリウム":   None,
        "カリウム":     None,
        "カルシウム":   None,
        "マグネシウム": None,
        "リン":         None,
        "鉄":           None,
        "亜鉛":         None,
        "ビタミンC":    None,
        "ビタミンD":    None,
        "ビタミンB12":  None,
    }

    ALIASES = {
        "食品番号":      ["食品番号"],
        "食品名":        ["食品名"],
        "エネルギー_kcal": ["エネルギー_kcal", "エネルギー（kcal）", "エネルギー(kcal)", "kcal", "ENERC_KCAL"],
        "水分":          ["水分", "WATER"],
        "たんぱく質":    ["たんぱく質", "タンパク質", "PROT-"],
        "脂質":          ["脂質", "FAT-"],
        "炭水化物":      ["炭水化物", "CHOCDF-", "CHOAVLDF-"],
        "食物繊維総量":  ["食物繊維総量", "食物繊維", "FIB-"],
        "灰分":          ["灰分", "ASH"],
        "ナトリウム":    ["ナトリウム", "NA"],
        "カリウム":      ["カリウム", "K"],
        "カルシウム":    ["カルシウム", "CA"],
        "マグネシウム":  ["マグネシウム", "MG"],
        "リン":          ["リン", "P"],
        "鉄":            ["鉄", "FE"],
        "亜鉛":          ["亜鉛", "ZN"],
        "ビタミンC":     ["ビタミンC", "ビタミン C", "ﾋﾞﾀﾐﾝC", "VITC"],
        "ビタミンD":     ["ビタミンD", "ビタミン D", "VITD"],
        "ビタミンB12":   ["ビタミンB12", "ビタミン B12", "VITB12"],
    }

    identifier_map = {
        "ENERC_KCAL": "エネルギー_kcal",
        "WATER": "水分",
        "PROT-": "たんぱく質",
        "FAT-": "脂質",
        "CHOCDF-": "炭水化物",
        "FIB-": "食物繊維総量",
        "ASH": "灰分",
        "NA": "ナトリウム",
        "K": "カリウム",
        "CA": "カルシウム",
        "MG": "マグネシウム",
        "P": "リン",
        "FE": "鉄",
        "ZN": "亜鉛",
        "VITC": "ビタミンC",
        "VITD": "ビタミンD",
        "VITB12": "ビタミンB12",
    }
    identifier_row = [str(c.value or "").strip() for c in ws[data_start_row - 1]]
    for ci, identifier in enumerate(identifier_row):
        canonical = identifier_map.get(identifier)
        if canonical and col_map[canonical] is None:
            col_map[canonical] = ci

    for ci, raw_name in enumerate(col_names):
        raw_norm = norm(raw_name)
        for canonical, aliases in ALIASES.items():
            if col_map[canonical] is None:
                for alias in aliases:
                    alias_norm = norm(alias)
                    if (
                        raw_norm == alias_norm
                        or (len(alias_norm) > 3 and alias_norm in raw_norm)
                    ):
                        col_map[canonical] = ci
                        break

    # 八訂2023の本表では列1=食品番号、列3=食品名、行12=成分識別子。
    if col_map["食品番号"] is None:
        col_map["食品番号"] = 1
    if col_map["食品名"] is None:
        col_map["食品名"] = 3

    return col_map


def convert(xlsx_path: Path, limit: int = 0) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl が必要です。")
        print("  python3 -m pip install openpyxl --target backend/.venv/lib/python3.9/site-packages")
        sys.exit(1)

    print(f"📂 読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # シートを選択（'本表' か最初のシート）
    sheet_name = next((s for s in wb.sheetnames if "本表" in s or "Table" in s), wb.sheetnames[0])
    ws = wb[sheet_name]
    print(f"📋 シート: {sheet_name}")

    # データ開始行を検出
    data_start_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        cells = [str(v or "").strip() for v in row[:4]]
        if any(re.match(r"^\d{5}$", cell) for cell in cells):
            data_start_row = row_idx
            break

    if data_start_row is None:
        print("ERROR: 食品番号（5桁数字）の行が見つかりません。")
        print("  別のシートが正しいかもしれません。シート一覧:", wb.sheetnames)
        sys.exit(1)

    print(f"📍 データ開始行: {data_start_row}")
    col_map = _build_col_map_from_headers(ws, data_start_row)

    # 列検出結果を表示
    detected = {k: v for k, v in col_map.items() if v is not None}
    print(f"🔍 検出列数: {len(detected)}/{len(col_map)}")
    missing = [k for k, v in col_map.items() if v is None]
    if missing:
        print(f"   ⚠️  未検出: {', '.join(missing)}")

    atoms = []
    skipped = 0

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        atom = _build_atom(row, col_map)
        if atom is None:
            skipped += 1
            continue
        atoms.append(atom)
        if limit and len(atoms) >= limit:
            break

    wb.close()
    return atoms


def main():
    parser = argparse.ArgumentParser(
        description="日本食品標準成分表（文部科学省）Excel を Atom JSON に変換",
        epilog=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("xlsx_path", nargs="?", help="MEXT Excel ファイルパス")
    parser.add_argument("legacy_limit", nargs="?", type=int, help="旧形式: 件数上限")
    parser.add_argument("--download-latest", action="store_true", help="MEXT 公式ページから最新 Excel を自動取得")
    parser.add_argument("--limit", type=int, default=0, help="変換する件数上限（テスト用）")
    parser.add_argument("--out", default=str(OUT_FILE), help="出力 JSON パス")
    args = parser.parse_args()

    if args.download_latest:
        xlsx_path = download_latest_excel()
    elif args.xlsx_path:
        xlsx_path = Path(args.xlsx_path)
    else:
        parser.print_help()
        sys.exit(0)

    if not xlsx_path.exists():
        print(f"ERROR: ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    limit = args.limit or args.legacy_limit or 0

    atoms = convert(xlsx_path, limit=limit)
    print(f"\n✅ 変換完了: {len(atoms)} Atom")

    # カテゴリ集計
    from collections import Counter
    cats = Counter(a["category"] for a in atoms)
    for cat, cnt in cats.most_common(10):
        print(f"   {cat}: {cnt}")

    out_file = Path(args.out)
    out_file.write_text(
        json.dumps(atoms, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n💾 保存: {out_file}  ({out_file.stat().st_size // 1024} KB)")
    print("\n次のステップ:")
    print("  python3 scripts/generate_docs.py  # progress.md を更新")


if __name__ == "__main__":
    main()
